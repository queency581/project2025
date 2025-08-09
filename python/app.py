import requests
from openpyxl import load_workbook
from datetime import datetime
import json

# === PATHS ===
CONFIG_PATH = r"C:\Users\mache\Desktop\quickbooks-oauth\New folder\quickbooks-oauth\config (1).json"
EXCEL_PATH = r"G:\My Drive\Trial balance sheet.xlsx"
SHEET_NAME = "Accrued non-current liabilities"

# === LOAD CONFIG ===
try:
    with open(CONFIG_PATH, "r") as f:
        config = json.load(f)
except Exception as e:
    print(f"❌ Failed to load config file: {e}")
    exit()

CLIENT_ID = config.get("client_id")
CLIENT_SECRET = config.get("client_secret")
REFRESH_TOKEN = config.get("refresh_token")

REALM_ID = "9341455066113633"
TOKEN_URL = "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer"

# === MONTH MAP ===
month_map = {
    "jan": "Jan",
    "feb": "Feb",
    "mar": "Mar",
    "apr": "April",
    "may": "May",
    "jun": "June",
    "jul": "July",
    "aug": "Aug",
    "sep": "Sep",
    "oct": "Oct",
    "nov": "Nov",
    "dec": "Dec"
}

# === DESCRIPTION MAP ===
description_map = {
    "change damaris chege": "Damaris Chege",
    "chnage simon maina": "Simon Maina",
    "change simon maina": "Simon Maina",
    "change quinter akoth": "Quinter Akoth",
    "overspend squeency": "Squeency",
    "dawn": "Dawn",
    "admin expenses": "Admin Expenses",
}

# === HELPERS ===
def normalize(text):
    return text.strip().lower() if text else ""

def find_staff_row(sheet, staff_name):
    staff_name_norm = normalize(staff_name)
    for row in range(2, sheet.max_row + 1):
        val = sheet.cell(row=row, column=1).value
        if val and normalize(val) == staff_name_norm:
            return row
    return None

def find_month_col(sheet, month_name):
    month_name_norm = normalize(month_name)
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        if val and normalize(val) == month_name_norm:
            return col
    return None

# === AUTH ===
def refresh_access_token():
    auth = (CLIENT_ID, CLIENT_SECRET)
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    payload = {
        "grant_type": "refresh_token",
        "refresh_token": REFRESH_TOKEN
    }

    response = requests.post(TOKEN_URL, auth=auth, headers=headers, data=payload)
    if response.status_code != 200:
        raise Exception(f"Token refresh failed: {response.status_code} {response.text}")
    
    tokens = response.json()
    config["refresh_token"] = tokens["refresh_token"]

    with open(CONFIG_PATH, "w") as f:
        json.dump(config, f, indent=4)

    print("✅ Saved new refresh token to config file")
    return tokens["access_token"]

# === API CALL ===
def query_journal_entries(access_token):
    url = f"https://quickbooks.api.intuit.com/v3/company/{REALM_ID}/query"
    query = "SELECT * FROM JournalEntry"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json",
        "Content-Type": "application/text"
    }

    response = requests.post(url, headers=headers, data=query)
    if response.status_code != 200:
        raise Exception(f"API call failed: {response.status_code} - {response.text}")
    
    return response.json()

# === MAIN ===
def main():
    print("🔑 Refreshing access token...")
    access_token = refresh_access_token()
    print("✅ Got new access token!")

    print("📦 Querying journal entries...")
    data = query_journal_entries(access_token)
    journal_entries = data.get("QueryResponse", {}).get("JournalEntry", [])

    if not journal_entries:
        print("⚠️ No Journal Entries found.")
        return

    print(f"✅ Found {len(journal_entries)} Journal Entries.")

    # === Open Excel ===
    try:
        book = load_workbook(EXCEL_PATH)
    except Exception as e:
        print(f"❌ Failed to load Excel file: {e}")
        return

    if SHEET_NAME not in book.sheetnames:
        print(f"❌ Sheet '{SHEET_NAME}' not found.")
        return

    sheet = book[SHEET_NAME]
    print(f"✅ Using sheet: '{SHEET_NAME}'")

    for je in journal_entries:
        txn_date_str = je.get("TxnDate")
        if not txn_date_str:
            continue
        txn_date = datetime.strptime(txn_date_str, "%Y-%m-%d")
        month_abbr = txn_date.strftime("%b").lower()
        excel_month = month_map.get(month_abbr)
        if not excel_month:
            continue

        for line in je.get("Line", []):
            description = line.get("Description", "").strip()
            desc_norm = normalize(description)
            staff_name = description_map.get(desc_norm, description).strip()
            amount = line.get("Amount", 0)

            staff_row = find_staff_row(sheet, staff_name)
            month_col = find_month_col(sheet, excel_month)

            if not staff_row or not month_col:
                print(f"⚠️ No match for '{staff_name}' or '{excel_month}'")
                continue

            sheet.cell(row=staff_row, column=month_col).value = amount
            print(f"✅ Updated {staff_name} | {excel_month} | Row {staff_row} Col {month_col} => {amount}")

    book.save(EXCEL_PATH)
    print(f"✅✅ Excel workbook updated: {EXCEL_PATH}")

# === ENTRY POINT ===
if __name__ == "__main__":
    main()
